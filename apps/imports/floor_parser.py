"""
Floor-level Excel parser.

Supported 1С formats:
  1. "Перемещение запасов" (.xls) — materials transfer to a floor
     Columns: col1=№, col3=Код, col8=Товары, col21=Количество, col26=Ед, col30=Цена
     total_amount = qty * price

  2. "Приход на склад" — warehouse receipt
     Columns: Номенклатура, Количество, Цена, Сумма

  3. Generic tabular — auto-detect columns

On re-import: match by item_code (if present) or by name.
"""

import re
from decimal import Decimal, InvalidOperation


def _to_decimal(value) -> Decimal:
    if value is None:
        return Decimal('0')
    s = str(value).strip().replace('\xa0', '').replace(' ', '').replace(',', '.')
    s = re.sub(r'[^\d.\-]', '', s)
    if not s or s in ('-', '.'):
        return Decimal('0')
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal('0')


def _str(v) -> str:
    if v is None:
        return ''
    return str(v).strip()


# ── Format detectors ──────────────────────────────────────────────────────────

def _is_peremeshchenie(ws, use_xlrd=False):
    """Detect 1С Перемещение запасов format."""
    try:
        if use_xlrd:
            v = str(ws.cell_value(1, 1)).lower()
        else:
            v = str(ws.cell(row=2, column=2).value or '').lower()
        return 'перемещение запасов' in v
    except Exception:
        return False


def _is_prihod(ws, use_xlrd=False):
    """Detect Приход на склад format."""
    try:
        if use_xlrd:
            row = [str(ws.cell_value(0, c)).lower() for c in range(min(ws.ncols, 5))]
        else:
            row = [str(c.value or '').lower() for c in list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0][:5]]
        return any('номенклатур' in v for v in row)
    except Exception:
        return False


# ── Parsers ───────────────────────────────────────────────────────────────────

def parse_peremeshchenie_xls(filepath) -> dict:
    """Parse 1С Перемещение запасов (.xls) format."""
    import xlrd
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    doc_title = _str(ws.cell_value(1, 1)) if ws.nrows > 1 else ''
    recipient = _str(ws.cell_value(5, 6)) if ws.nrows > 5 else ''

    items = []
    for r in range(ws.nrows):
        # Row №: column 1 should be a number
        num_raw = ws.cell_value(r, 1)
        if not num_raw:
            continue
        try:
            float(num_raw)
        except (ValueError, TypeError):
            continue  # skip header/meta rows

        code = _str(ws.cell_value(r, 3))
        name = _str(ws.cell_value(r, 8))
        qty_raw   = ws.cell_value(r, 21) if ws.ncols > 21 else None
        unit      = _str(ws.cell_value(r, 26)) if ws.ncols > 26 else ''
        price_raw = ws.cell_value(r, 30) if ws.ncols > 30 else None

        if not name:
            continue

        qty   = _to_decimal(qty_raw)
        price = _to_decimal(price_raw)
        total = qty * price

        items.append({
            'item_code': code,
            'name': name,
            'unit': unit,
            'quantity': qty,
            'unit_price': price,
            'total_amount': total,
        })

    return {
        'doc_title': doc_title,
        'recipient': recipient,
        'items': items,
        'format': 'peremeshchenie',
    }


def parse_prihod(filepath) -> dict:
    """Parse Приход на склад format (Номенклатура, Количество, Цена, Сумма)."""
    import xlrd, openpyxl

    ext = filepath.rsplit('.', 1)[-1].lower()

    if ext == 'xls':
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    else:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = list(wb.worksheets)[0]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()

    # Find header row
    header_idx = 0
    col_name = col_qty = col_price = col_total = None
    for i, row in enumerate(rows[:10]):
        row_lower = [str(v or '').lower() for v in row]
        if any('номенклатур' in v for v in row_lower):
            header_idx = i
            for j, v in enumerate(row_lower):
                if 'номенклатур' in v: col_name = j
                elif 'количеств' in v: col_qty = j
                elif 'цен' in v and col_price is None: col_price = j
                elif 'сумм' in v: col_total = j
            break

    items = []
    current_name = None
    for r in range(header_idx + 1, len(rows)):
        row = rows[r]
        if not any(v for v in row if v is not None):
            continue

        name_val = _str(row[col_name]) if col_name is not None and col_name < len(row) else ''
        qty_val = row[col_qty] if col_qty is not None and col_qty < len(row) else None
        price_val = row[col_price] if col_price is not None and col_price < len(row) else None
        total_val = row[col_total] if col_total is not None and col_total < len(row) else None

        # Hierarchical: name without qty = category header
        if name_val and not qty_val:
            current_name = name_val
            continue

        name = name_val or current_name or ''
        if not name:
            continue

        qty = _to_decimal(qty_val)
        price = _to_decimal(price_val)
        total = _to_decimal(total_val)
        if total == 0 and qty > 0 and price > 0:
            total = qty * price

        items.append({
            'item_code': '',
            'name': name,
            'unit': '',
            'quantity': qty,
            'unit_price': price,
            'total_amount': total,
        })

    return {
        'doc_title': 'Приход на склад',
        'recipient': '',
        'items': items,
        'format': 'prihod',
    }


def parse_generic(filepath) -> dict:
    """
    Generic parser: auto-detect name/qty/price/total columns.
    Works for both .xls and .xlsx.
    """
    import xlrd, openpyxl

    ext = filepath.rsplit('.', 1)[-1].lower()

    if ext == 'xls':
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    else:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = list(wb.worksheets)[0]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()

    # Find header row
    header_idx = 0
    col_code = col_name = col_unit = col_qty = col_price = col_total = None

    for i, row in enumerate(rows[:15]):
        row_lower = [str(v or '').lower().strip() for v in row]
        str_count = sum(1 for v in row_lower if v and not v.replace('.', '').replace('-', '').isnumeric())
        if str_count < 2:
            continue
        header_idx = i
        for j, v in enumerate(row_lower):
            if v in ('код', 'code', 'артикул') and col_code is None: col_code = j
            if any(k in v for k in ['наименован', 'товар', 'материал', 'name', 'номенклатур']) and col_name is None: col_name = j
            if any(k in v for k in ['ед', 'unit', 'изм']) and col_unit is None: col_unit = j
            if any(k in v for k in ['количеств', 'qty', 'кол']) and col_qty is None: col_qty = j
            if any(k in v for k in ['цена', 'price', 'стоимость ед']) and col_price is None: col_price = j
            if any(k in v for k in ['сумм', 'total', 'итог', 'стоимость']) and col_total is None: col_total = j
        if col_name is not None or col_qty is not None:
            break

    items = []
    for r in range(header_idx + 1, len(rows)):
        row = rows[r]
        if not any(v for v in row if v is not None):
            continue

        def gcol(idx):
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        name = _str(gcol(col_name)) if col_name is not None else ''
        if not name:
            continue

        code = _str(gcol(col_code)) if col_code is not None else ''
        unit = _str(gcol(col_unit)) if col_unit is not None else ''
        qty = _to_decimal(gcol(col_qty))
        price = _to_decimal(gcol(col_price))
        total = _to_decimal(gcol(col_total))
        if total == 0 and qty > 0 and price > 0:
            total = qty * price

        if name.lower() in ('итого', 'всего', 'total') or (not qty and not total):
            continue

        items.append({
            'item_code': code,
            'name': name,
            'unit': unit,
            'quantity': qty,
            'unit_price': price,
            'total_amount': total,
        })

    return {
        'doc_title': '',
        'recipient': '',
        'items': items,
        'format': 'generic',
    }


def parse_floor_excel(filepath) -> dict:
    """Main entry point: auto-detect format and parse."""
    ext = filepath.rsplit('.', 1)[-1].lower()

    if ext == 'xls':
        try:
            import xlrd
            wb = xlrd.open_workbook(filepath)
            ws = wb.sheet_by_index(0)
            if _is_peremeshchenie(ws, use_xlrd=True):
                return parse_peremeshchenie_xls(filepath)
            if _is_prihod(ws, use_xlrd=True):
                return parse_prihod(filepath)
        except Exception:
            pass
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = list(wb.worksheets)[0]
            if _is_prihod(ws, use_xlrd=False):
                wb.close()
                return parse_prihod(filepath)
            wb.close()
        except Exception:
            pass

    return parse_generic(filepath)


def apply_floor_import(floor, parsed_data, import_mode='replace', excel_import=None):
    """
    Apply parsed floor expense data to the database.

    import_mode:
        'replace' — update quantity/price if code/name matches; create if new
        'add'     — add quantities to existing records; create if new

    Returns dict with stats: created, updated, skipped.
    """
    from apps.projects.models import FloorExpense

    stats = {'created': 0, 'updated': 0, 'errors': 0}
    items = parsed_data.get('items', [])

    for item in items:
        try:
            code = item.get('item_code', '').strip()
            name = item.get('name', '').strip()
            if not name:
                continue

            qty = item.get('quantity', Decimal('0'))
            price = item.get('unit_price', Decimal('0'))
            total = item.get('total_amount', Decimal('0'))

            # Match existing expense
            existing = None
            if code:
                existing = floor.expenses.filter(item_code=code).first()
            if not existing:
                existing = floor.expenses.filter(name__iexact=name).first()

            if existing:
                if import_mode == 'add':
                    existing.quantity += qty
                    if price > 0:
                        existing.unit_price = price
                    existing.total_amount = existing.quantity * existing.unit_price if existing.unit_price else existing.total_amount + total
                else:  # replace
                    existing.quantity = qty
                    if price > 0:
                        existing.unit_price = price
                    existing.total_amount = total if total else (qty * price if price else Decimal('0'))
                if code:
                    existing.item_code = code
                existing.source_import = excel_import
                existing.save()
                stats['updated'] += 1
            else:
                FloorExpense.objects.create(
                    floor=floor,
                    item_code=code,
                    name=name,
                    unit=item.get('unit', ''),
                    quantity=qty,
                    unit_price=price,
                    total_amount=total if total else (qty * price),
                    source_import=excel_import,
                )
                stats['created'] += 1
        except Exception as e:
            stats['errors'] += 1

    return stats
