"""
Excel parser for 1С exports and custom templates.

Supported formats:
  1. Custom template: columns ЖК | Блок | Этап | План | Факт
  2. 1С ОБОРОТНО-САЛЬДОВАЯ ВЕДОМОСТЬ style
  3. 1С Отчёт по затратам with subcontos
  4. Any Excel with recognizable column headers
"""

import re
from decimal import Decimal, InvalidOperation


COLUMN_KEYWORDS = {
    'complex': ['жк', 'жилой комплекс', 'объект строительства', 'объект', 'комплекс', 'residential'],
    'block': ['блок', 'block', 'секция', 'подъезд'],
    'stage': ['этап', 'stage', 'статья затрат', 'статья', 'работа', 'вид работ', 'наименование работ'],
    'planned': ['план', 'planned', 'смета', 'сметная стоимость', 'план расходов', 'budget'],
    'actual': ['факт', 'actual', 'расход', 'затраты', 'фактические затраты', 'факт расходов', 'сумма'],
    'date': ['дата', 'date', 'период'],
    'note': ['примечание', 'note', 'комментарий', 'описание'],
}


def _to_decimal(value) -> Decimal:
    if value is None:
        return Decimal('0')
    s = str(value).strip().replace(' ', '').replace('\xa0', '').replace(',', '.')
    s = re.sub(r'[^\d.\-]', '', s)
    if not s or s in ('-', '.'):
        return Decimal('0')
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal('0')


def _find_header_row(rows):
    """Find the row index that looks like a header (has multiple string cells)."""
    for i, row in enumerate(rows[:20]):
        str_cells = [c for c in row if isinstance(c, str) and len(str(c).strip()) > 1]
        if len(str_cells) >= 2:
            return i
    return 0


def _detect_column_mapping(header_row):
    """Map column keywords to column indexes based on header values."""
    mapping = {}
    for col_idx, cell in enumerate(header_row):
        cell_lower = str(cell).lower().strip()
        for field, keywords in COLUMN_KEYWORDS.items():
            if field not in mapping and any(kw in cell_lower for kw in keywords):
                mapping[field] = col_idx
                break
    return mapping


def _get_cell(row, col_idx):
    if col_idx is None or col_idx >= len(row):
        return None
    val = row[col_idx]
    if val is None or (isinstance(val, float) and str(val) == 'nan'):
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ('none', 'nan', '-') else None


class ExcelParser:
    def __init__(self, filepath, column_mapping=None):
        self.filepath = filepath
        self.forced_mapping = column_mapping or {}
        self.detected_mapping = {}
        self.errors = []
        self.warnings = []

    def parse(self):
        """
        Parse the Excel file and return list of dicts:
        [{'complex_name', 'block_name', 'stage_name', 'planned', 'actual', 'row_num'}, ...]
        """
        import openpyxl

        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        ws = wb.active

        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))

        wb.close()

        if not all_rows:
            self.errors.append('Файл пустой')
            return []

        header_idx = _find_header_row(all_rows)
        header = all_rows[header_idx]

        if self.forced_mapping:
            mapping = self.forced_mapping
        else:
            mapping = _detect_column_mapping(header)
            self.detected_mapping = mapping

        results = []
        current_complex = None
        current_block = None

        for row_num, row in enumerate(all_rows[header_idx + 1:], start=header_idx + 2):
            # Skip fully empty rows
            non_empty = [c for c in row if c is not None and str(c).strip() not in ('', 'None', 'nan')]
            if not non_empty:
                continue

            complex_name = _get_cell(row, mapping.get('complex'))
            block_name = _get_cell(row, mapping.get('block'))
            stage_name = _get_cell(row, mapping.get('stage'))
            planned_raw = _get_cell(row, mapping.get('planned'))
            actual_raw = _get_cell(row, mapping.get('actual'))

            # Inherit context for hierarchical 1С-style exports
            if complex_name:
                current_complex = complex_name
            if block_name:
                current_block = block_name

            effective_complex = complex_name or current_complex
            effective_block = block_name or current_block

            if not stage_name and not effective_complex:
                continue

            # Skip subtotal/total rows
            if stage_name and any(
                kw in stage_name.lower()
                for kw in ['итого', 'всего', 'total', 'subtotal', 'итог']
            ):
                continue

            planned = _to_decimal(planned_raw)
            actual = _to_decimal(actual_raw)

            if planned == 0 and actual == 0 and not stage_name:
                continue

            results.append({
                'complex_name': effective_complex,
                'block_name': effective_block,
                'stage_name': stage_name,
                'planned': planned,
                'actual': actual,
                'row_num': row_num,
            })

        return results

    def get_preview(self, max_rows=20):
        """Return raw header + first N rows for preview."""
        import openpyxl
        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        ws = wb.active

        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))
            if len(all_rows) > max_rows + 20:
                break

        wb.close()

        header_idx = _find_header_row(all_rows)
        headers = [str(c) if c is not None else '' for c in all_rows[header_idx]]
        data_rows = []
        for row in all_rows[header_idx + 1: header_idx + 1 + max_rows]:
            data_rows.append([str(c) if c is not None else '' for c in row])

        mapping = _detect_column_mapping(all_rows[header_idx])

        return {
            'headers': headers,
            'rows': data_rows,
            'detected_mapping': mapping,
        }
